import os
import matplotlib
matplotlib.use('Agg')  # Безопасный режим для потоков

import matplotlib.pyplot as plt
import seaborn as sns

import plotly.express as px
import plotly.io as pio

import numpy as np
import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

import smtplib
import mimetypes
from datetime import datetime
from email.message import EmailMessage
import ssl

class DataReport:
    """
    Генерация отчётов с визуализацией и отправкой по email.
    """

    def __init__(self, df: pd.DataFrame, report_name: str = "report"):
        self.df = df.copy()
        self.report_name = report_name
        self.created_at = datetime.now()

    def generate_key_metrics(self) -> dict:
        numeric = self.df.select_dtypes(include=[np.number])
        metrics = {
            "generated_at": self.created_at.isoformat(timespec="seconds"),
            "rows": int(self.df.shape[0]),
            "columns": int(self.df.shape[1]),
            "missing_values_total": int(self.df.isna().sum().sum()),
            "missing_values_by_column": self.df.isna().sum().to_dict(),
            "numeric_columns": list(numeric.columns),
        }

        if not numeric.empty:
            metrics["numeric_summary"] = {
                col: {
                    "mean": float(numeric[col].mean()),
                    "median": float(numeric[col].median()),
                    "std": float(numeric[col].std()),
                    "min": float(numeric[col].min()),
                    "max": float(numeric[col].max()),
                }
                for col in numeric.columns
            }
            metrics["correlation"] = numeric.corr(numeric_only=True).to_dict()

        return metrics

    def save_matplotlib_seaborn_plots(self, out_dir: str = "reports/figures", max_numeric_cols: int = 6) -> dict:
        os.makedirs(out_dir, exist_ok=True)
        paths = {}

        numeric = self.df.select_dtypes(include=[np.number])
        cols = list(numeric.columns)[:max_numeric_cols]

        for col in cols:
            plt.figure(figsize=(7, 4))
            sns.histplot(numeric[col].dropna(), kde=True)
            plt.title(f"Histogram: {col}")
            plt.tight_layout()
            p = os.path.join(out_dir, f"{self.report_name}_hist_{col}.png")
            plt.savefig(p, dpi=150)
            plt.close()
            paths[f"hist_{col}"] = p

        if numeric.shape[1] >= 2:
            plt.figure(figsize=(8, 6))
            corr = numeric.corr(numeric_only=True)
            sns.heatmap(corr, cmap="coolwarm", center=0, square=False)
            plt.title("Correlation heatmap")
            plt.tight_layout()
            p = os.path.join(out_dir, f"{self.report_name}_corr_heatmap.png")
            plt.savefig(p, dpi=150)
            plt.close()
            paths["corr_heatmap"] = p

        return paths

    def save_plotly_interactive(self, out_dir: str = "reports/interactive", max_numeric_cols: int = 4) -> dict:
        os.makedirs(out_dir, exist_ok=True)
        paths = {}

        numeric = self.df.select_dtypes(include=[np.number])
        cols = list(numeric.columns)

        if len(cols) >= 2:
            fig = px.scatter(self.df, x=cols[0], y=cols[1], title=f"Scatter: {cols[0]} vs {cols[1]}")
            p = os.path.join(out_dir, f"{self.report_name}_scatter_{cols[0]}_{cols[1]}.html")
            pio.write_html(fig, file=p, auto_open=False, include_plotlyjs="cdn")
            paths["scatter"] = p

        for col in cols[:max_numeric_cols]:
            fig = px.box(self.df, y=col, title=f"Boxplot: {col}")
            p = os.path.join(out_dir, f"{self.report_name}_box_{col}.html")
            pio.write_html(fig, file=p, auto_open=False, include_plotlyjs="cdn")
            paths[f"box_{col}"] = p

        return paths

    def export_pdf(self, pdf_path: str = None, include_figures: bool = True, figures_dir: str = "reports/figures") -> str:
        os.makedirs(os.path.dirname(pdf_path) if pdf_path else "reports", exist_ok=True)
        if pdf_path is None:
            pdf_path = os.path.join("reports", f"{self.report_name}.pdf")

        metrics = self.generate_key_metrics()

        c = canvas.Canvas(pdf_path, pagesize=A4)
        width, height = A4
        y = height - 2 * cm

        c.setFont("Helvetica-Bold", 14)
        c.drawString(2 * cm, y, f"Data Report: {self.report_name}")
        y -= 1 * cm

        c.setFont("Helvetica", 10)
        lines = [
            f"Generated at: {metrics['generated_at']}",
            f"Rows: {metrics['rows']}",
            f"Columns: {metrics['columns']}",
            f"Missing values (total): {metrics['missing_values_total']}",
        ]
        for line in lines:
            c.drawString(2 * cm, y, line)
            y -= 0.6 * cm

        y -= 0.4 * cm
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2 * cm, y, "Missing values by column (top 15):")
        y -= 0.7 * cm
        c.setFont("Helvetica", 9)

        missing_sorted = sorted(metrics["missing_values_by_column"].items(), key=lambda x: x[1], reverse=True)[:15]
        for col, cnt in missing_sorted:
            c.drawString(2 * cm, y, f"- {col}: {cnt}")
            y -= 0.45 * cm
            if y < 2 * cm:
                c.showPage()
                y = height - 2 * cm
                c.setFont("Helvetica", 9)

        if "numeric_summary" in metrics and metrics["numeric_summary"]:
            y -= 0.4 * cm
            c.setFont("Helvetica-Bold", 11)
            c.drawString(2 * cm, y, "Numeric summary (top 10 columns):")
            y -= 0.7 * cm
            c.setFont("Helvetica", 9)

            for col, s in list(metrics["numeric_summary"].items())[:10]:
                c.drawString(2 * cm, y, f"- {col}: mean={s['mean']:.4f}, median={s['median']:.4f}, std={s['std']:.4f}")
                y -= 0.45 * cm
                if y < 2 * cm:
                    c.showPage()
                    y = height - 2 * cm
                    c.setFont("Helvetica", 9)

        if include_figures:
            y -= 0.4 * cm
            c.setFont("Helvetica-Bold", 11)
            c.drawString(2 * cm, y, f"Figures saved in: {figures_dir}")
            y -= 0.7 * cm
            c.setFont("Helvetica", 9)

            if os.path.isdir(figures_dir):
                fig_files = sorted([f for f in os.listdir(figures_dir) if f.lower().endswith(".png")])[:200]
                for f in fig_files:
                    c.drawString(2 * cm, y, f"- {f}")
                    y -= 0.45 * cm
                    if y < 2 * cm:
                        c.showPage()
                        y = height - 2 * cm
                        c.setFont("Helvetica", 9)

        c.save()
        return pdf_path

    def export_excel(self, xlsx_path: str = None, include_sample_rows: int = 50) -> str:
        os.makedirs(os.path.dirname(xlsx_path) if xlsx_path else "reports", exist_ok=True)
        if xlsx_path is None:
            xlsx_path = os.path.join("reports", f"{self.report_name}.xlsx")

        metrics = self.generate_key_metrics()
        wb = Workbook()

        ws = wb.active
        ws.title = "metrics"
        ws.append(["metric", "value"])
        for k, v in [
            ("generated_at", metrics["generated_at"]),
            ("rows", metrics["rows"]),
            ("columns", metrics["columns"]),
            ("missing_values_total", metrics["missing_values_total"]),
        ]:
            ws.append([k, v])

        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40

        ws2 = wb.create_sheet("missing")
        ws2.append(["column", "missing_count"])
        for col, cnt in sorted(metrics["missing_values_by_column"].items(), key=lambda x: x[1], reverse=True):
            ws2.append([col, cnt])
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)
        ws2.column_dimensions["A"].width = 35
        ws2.column_dimensions["B"].width = 16

        ws3 = wb.create_sheet("numeric_summary")
        ws3.append(["column", "mean", "median", "std", "min", "max"])
        if "numeric_summary" in metrics:
            for col, s in metrics["numeric_summary"].items():
                ws3.append([col, s["mean"], s["median"], s["std"], s["min"], s["max"]])

        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions["A"].width = 28

        ws4 = wb.create_sheet("data_sample")
        sample = self.df.head(include_sample_rows)
        for r in dataframe_to_rows(sample, index=False, header=True):
            ws4.append(r)
        for cell in ws4[1]:
            cell.font = Font(bold=True)

        wb.save(xlsx_path)
        return xlsx_path

    @staticmethod
    def send_email_smtp(
            smtp_host: str,
            smtp_port: int,
            username: str,
            password: str,
            sender: str,
            recipients: list,
            subject: str,
            body: str,
            attachments: list = None,
            use_tls: bool = True,
    ):
        """
        Отправка email через SMTP с вложениями.
        """
        msg = EmailMessage()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        msg.set_content(body)

        attachments = attachments or []
        for path in attachments:
            if not path or not os.path.isfile(path):
                continue

            ctype, encoding = mimetypes.guess_type(path)
            if ctype is None or encoding is not None:
                ctype = "application/octet-stream"
            maintype, subtype = ctype.split("/", 1)

            with open(path, "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype=maintype,
                    subtype=subtype,
                    filename=os.path.basename(path),
                )

        try:
            if use_tls:
                context = ssl.create_default_context()
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    server.starttls(context=context)
                    if username and password:
                        server.login(username, password)
                    server.send_message(msg)
            else:
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    if username and password:
                        server.login(username, password)
                    server.send_message(msg)
        except smtplib.SMTPAuthenticationError:
            raise Exception("❌ Ошибка аутентификации: неверный email или пароль")
        except smtplib.SMTPRecipientsRefused:
            raise Exception("❌ Один или несколько получателей отклонены сервером")
        except Exception as e:
            raise Exception(f"❌ Не удалось отправить письмо: {e}")

    @staticmethod
    def _collect_report_files(out_dir: str) -> list:
        files = []
        if not out_dir or not os.path.isdir(out_dir):
            return files

        for root, _, filenames in os.walk(out_dir):
            for name in filenames:
                p = os.path.join(root, name)
                if os.path.isfile(p):
                    files.append(p)

        files.sort()
        return files

    @classmethod
    def send_full_report_smtp(
        cls,
        artifacts: dict,
        smtp_host: str,
        smtp_port: int,
        username: str,
        password: str,
        sender: str,
        recipients: list,
        subject: str,
        body: str,
        use_tls: bool = True,
        include_all_out_dir_files: bool = True,
    ) -> list:
        attachments = []

        def add(p: str):
            if p and os.path.isfile(p) and p not in attachments:
                attachments.append(p)

        out_dir = artifacts.get("out_dir")

        if include_all_out_dir_files and out_dir:
            attachments = cls._collect_report_files(out_dir)
        else:
            add(artifacts.get("pdf"))
            add(artifacts.get("excel"))

            sp = artifacts.get("static_plots") or {}
            for _, p in sp.items():
                add(p)

            ip = artifacts.get("interactive_plots") or {}
            for _, p in ip.items():
                add(p)

        cls.send_email_smtp(
            smtp_host=smtp_host,
            smtp_port=smtp_port,
            username=username,
            password=password,
            sender=sender,
            recipients=recipients,
            subject=subject,
            body=body,
            attachments=attachments,
            use_tls=use_tls,
        )
        return attachments

    def build_full_report(
        self,
        out_dir: str = "reports",
        make_pdf: bool = True,
        make_excel: bool = True,
        make_static_plots: bool = True,
        make_interactive_plots: bool = True,
    ) -> dict:
        os.makedirs(out_dir, exist_ok=True)

        artifacts = {"out_dir": out_dir}
        artifacts["metrics"] = self.generate_key_metrics()

        if make_static_plots:
            artifacts["static_plots"] = self.save_matplotlib_seaborn_plots(
                out_dir=os.path.join(out_dir, "figures")
            )
        if make_interactive_plots:
            artifacts["interactive_plots"] = self.save_plotly_interactive(
                out_dir=os.path.join(out_dir, "interactive")
            )
        if make_pdf:
            artifacts["pdf"] = self.export_pdf(
                pdf_path=os.path.join(out_dir, f"{self.report_name}.pdf"),
                include_figures=make_static_plots,
                figures_dir=os.path.join(out_dir, "figures"),
            )
        if make_excel:
            artifacts["excel"] = self.export_excel(
                xlsx_path=os.path.join(out_dir, f"{self.report_name}.xlsx")
            )

        return artifacts